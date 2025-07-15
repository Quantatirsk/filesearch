"""
XLSX parser using python-calamine.

Following the technical report's recommendation for python-calamine
as the fastest XLSX parsing library with Rust-based implementation.
"""

from typing import Optional
from .base_parser import BaseParser, ParserFactory


class XLSXParser(BaseParser):
    """
    High-performance XLSX parser using python-calamine.
    
    As recommended in the technical report, python-calamine is 10-80x faster
    than openpyxl thanks to its Rust-based Calamine library implementation.
    """
    
    def parse(self, file_path: str) -> Optional[str]:
        """
        Parse an XLSX file and extract text content.
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            Extracted text content or None if parsing fails
        """
        try:
            import pandas as pd
            
            # Use calamine engine for maximum performance
            # This is 10-80x faster than the default openpyxl engine
            df = pd.read_excel(file_path, engine='calamine', sheet_name=None)
            
            text_content = []
            
            # Process all sheets
            for sheet_name, sheet_data in df.items():
                # Add sheet name as header
                text_content.append(f"=== {sheet_name} ===")
                
                # Convert DataFrame to string representation
                sheet_text = sheet_data.to_string(index=False, na_rep='')
                text_content.append(sheet_text)
                text_content.append("")  # Empty line between sheets
            
            return '\n'.join(text_content)
            
        except ImportError:
            print("python-calamine is not installed. Please install it with: pip install python-calamine")
            return None
        except Exception as e:
            print(f"Error parsing XLSX file {file_path}: {e}")
            return None
    
    def get_supported_extensions(self) -> list:
        """Get supported file extensions."""
        return ['.xlsx']


class XLSXParserFallback(BaseParser):
    """
    Fallback XLSX parser using openpyxl.
    
    This is a fallback option if python-calamine is not available,
    though it will be significantly slower.
    """
    
    def parse(self, file_path: str) -> Optional[str]:
        """
        Parse an XLSX file using openpyxl.
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            Extracted text content or None if parsing fails
        """
        try:
            import openpyxl
            
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            text_content = []
            
            # Process all sheets
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Add sheet name as header
                text_content.append(f"=== {sheet_name} ===")
                
                # Extract text from all cells
                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None:
                            row_text.append(str(cell))
                        else:
                            row_text.append("")
                    
                    if any(cell.strip() for cell in row_text):  # Skip empty rows
                        text_content.append('\t'.join(row_text))
                
                text_content.append("")  # Empty line between sheets
            
            workbook.close()
            return '\n'.join(text_content)
            
        except ImportError:
            print("openpyxl is not installed. Please install it with: pip install openpyxl")
            return None
        except Exception as e:
            print(f"Error parsing XLSX file {file_path}: {e}")
            return None
    
    def get_supported_extensions(self) -> list:
        """Get supported file extensions."""
        return ['.xlsx']


# Register the primary XLSX parser
ParserFactory.register_parser(XLSXParser)